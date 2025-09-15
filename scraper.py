import requests
import pandas as pd
from bs4 import BeautifulSoup
import time
import os
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment


load_dotenv()
APIFY_TOKEN = "apify_api_Ed2WeEoSp0Y1RfqjmzopONZsLSVQ9h4yIgEr"
ACTOR_ID = ("ACTOR_ID")

if not APIFY_TOKEN or not ACTOR_ID:
    print("‚ùå ERROR: Missing APIFY_TOKEN or ACTOR_ID in .env file")
    exit()


job_title = input("Enter Job Title: ")
max_jobs = 50   
url = f"https://api.apify.com/v2/acts/{ACTOR_ID}/runs?token={APIFY_TOKEN}"
payload = {
    "startUrls": [{"url": f"https://www.indeed.com/jobs?q={job_title}"}],
    "maxResults": 100
}

response = requests.post(url, json=payload)
run_data = response.json()


if "data" not in run_data:
    print("‚ùå API Error:", run_data)
    exit()

run_id = run_data["data"]["id"]


status_url = f"https://api.apify.com/v2/acts/{ACTOR_ID}/runs/{run_id}?token={APIFY_TOKEN}"

print("\n‚è≥ Fetching jobs... Please wait.\n")

while True:
    status_response = requests.get(status_url).json()
    status = status_response["data"]["status"]
    if status in ["SUCCEEDED", "FAILED", "ABORTED"]:
        break
    print("...still fetching, please wait...")
    time.sleep(5)

if status != "SUCCEEDED":
    print(f"‚ùå Actor run failed with status: {status}")
    exit()


dataset_id = status_response["data"]["defaultDatasetId"]
dataset_url = f"https://api.apify.com/v2/datasets/{dataset_id}/items?format=json&token={APIFY_TOKEN}"
data = requests.get(dataset_url).json()


cleaned = []
skills_keywords = ["python", "java", "sql", "excel", "machine learning", "aws", "django", "flask"]

for job in data:
    if len(cleaned) >= max_jobs:
        break

    desc_html = job.get("descriptionHTML", "")
    soup = BeautifulSoup(desc_html, "html.parser")
    desc_text = soup.get_text(" ", strip=True)

    found_skills = [s for s in skills_keywords if s.lower() in desc_text.lower()]

    cleaned.append({
        "Job ID": job.get("id"),
        "Job Title": job.get("positionName"),
        "Company": job.get("company"),
        "Location": job.get("location"),
        "Remote": "Yes" if "remote" in (job.get("location") or "").lower() else "No",
        "Salary": job.get("salary"),
        "Job Type": ", ".join(job.get("jobType", [])) if job.get("jobType") else "",
        "Rating": job.get("rating"),
        "Reviews": job.get("reviewsCount"),
        "Posted": job.get("postedAt"),
        "Benefits": ", ".join(job.get("benefits", [])) if job.get("benefits") else "",
        "Skills (detected)": ", ".join(found_skills),
        "Apply Link": job.get("externalApplyLink") or job.get("url"),
        "Description": desc_text[:500] + "..." if len(desc_text) > 500 else desc_text
    })


if not cleaned:
    print("‚ö†Ô∏è No jobs found. Try another search.")
    exit()

df = pd.DataFrame(cleaned)
df.drop_duplicates(subset=["Job Title", "Company", "Location"], inplace=True)

if "Posted" in df.columns:
    df = df.sort_values(by="Posted", ascending=False)

file_name = f"{job_title}_cleaned_jobs.xlsx"
df.to_excel(file_name, index=False)
df.to_csv(f"{job_title}_cleaned_jobs.csv", index=False)

# Excel formatting
wb = load_workbook(file_name)
ws = wb.active

header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

for col in ws.columns:
    max_length = 0
    col_name = col[0].column_letter
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_name].width = max_length + 2

wb.save(file_name)

print(f"\n‚úÖ Saved {len(df)} cleaned and formatted jobs to {file_name} & CSV file.")
print(f"üìä Total Jobs Collected: {len(df)} | Unique Companies: {df['Company'].nunique()}")
